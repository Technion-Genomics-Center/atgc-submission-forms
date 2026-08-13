# -*- coding: utf-8 -*-
"""Survey every submission workbook and report its real field structure.

The submission forms drifted apart over years of copy-paste, so the schema has
to be derived from the files rather than assumed. This reads all of them and
prints, per application: the header fields, the choice fields (with the
`Setting` column each dropdown points at), the sample-table columns, and the
vocabularies the `Setting` sheet carries.

Read-only. Writes nothing outside SubmissionForm/.

    python tools/survey_workbooks.py            # human-readable
    python tools/survey_workbooks.py --json     # machine-readable, for build.py
"""

from __future__ import annotations

import glob
import json
import re
import sys
import warnings
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore", message=".*Data Validation extension.*")

WEBSITE = Path(r"Y:\LAB\TGC_website\ATGC_website_2025\Applications")
FORMS = Path(r"Y:\LAB\Forms\Submission_forms")

# The two workbooks that were never published to the website (doc 05 §2).
UNPUBLISHED = ["sc10X-electronic_2025.xlsx", "Spatial-electronic_2025.xlsx"]

# Not built — doc 05 §12.4. Surveyed anyway so the report is complete and the
# decision is visible rather than looking like an oversight.
NOT_BUILT = {
    "CELseq-electronic 2025.xlsx": "service withdrawn",
    "RNAseq-extraction-electronic 2025.xlsx": "merged into RNA-seq",
    "Metagenomics-extraction-electronicV2_2025.xlsx": "merged into 16S/18S + shotgun",
}

CELL_RE = re.compile(r"\$?([A-Z]{1,2})\$?(\d+)")


def workbooks():
    """Every submission workbook, website set first."""
    found = sorted(glob.glob(str(WEBSITE / "**" / "*electronic*.xlsx"), recursive=True))
    found += [str(FORMS / name) for name in UNPUBLISHED]
    return found


def sample_table(ws):
    """Locate the sample table and return (header_row, [(coord, label), ...]).

    Every form anchors its table on a `Sample ID` cell, so find that and walk
    right until the labels stop.
    """
    for row in ws.iter_rows(max_row=min(ws.max_row, 60)):
        for cell in row:
            if str(cell.value).strip().lower() == "sample id":
                cols = []
                for c in ws[cell.row][cell.column - 1:]:
                    v = c.value
                    if v is None or not str(v).strip():
                        break
                    cols.append((c.coordinate, str(v).strip()))
                return cell.row, cols
    return None, []


def validation_targets(ws):
    """Map each validated cell to the `Setting` range it draws from.

    This is what distinguishes a choice field from a free-text one; guessing
    from the label would misclassify both ways.
    """
    out = {}
    for dv in ws.data_validations.dataValidation:
        src = (dv.formula1 or "").strip()
        for rng in dv.sqref.ranges:
            block = ws[str(rng)]
            # A single-cell range comes back as a bare Cell, not a row tuple.
            if not isinstance(block, tuple):
                block = ((block,),)
            for row in block:
                row = row if isinstance(row, tuple) else (row,)
                for cell in row:
                    out[cell.coordinate] = src
    return out


def setting_columns(wb):
    """The `Setting` sheet's vocabularies, keyed by their header.

    A single column can hold SEVERAL vocabularies stacked with a blank row
    between them — column A is typically `YesNo` (Yes, No), a gap, then `Run#`
    (0.5, 1, 2, Other). Reading the column whole would splice them into one list
    and put "Run#" in the Yes/No dropdown, so split on blank cells and treat the
    first cell of each block as its header.
    """
    if "Setting" not in wb.sheetnames:
        return {}
    ws = wb["Setting"]
    vocab = {}
    for col in ws.iter_cols():
        block = []
        for cell in col:
            text = "" if cell.value is None else str(cell.value).strip()
            if text:
                block.append(text)
                continue
            if len(block) > 1:
                vocab[block[0]] = block[1:]
            block = []
        if len(block) > 1:
            vocab[block[0]] = block[1:]
    return vocab


def label_for(ws, coord, vocab_values=frozenset()):
    """The nearest label to the left of a control cell, on the same row.

    Forms pre-fill their dropdowns, so the cell immediately left of a control is
    often another control holding a VALUE ("Nextseq", "P1 300 cycles"), not a
    label. Skip anything that is itself a `Setting` value and keep walking.
    """
    m = CELL_RE.match(coord)
    if not m:
        return ""
    col = openpyxl.utils.column_index_from_string(m.group(1))
    row = int(m.group(2))
    for c in range(col - 1, 0, -1):
        v = ws.cell(row=row, column=c).value
        if v is None:
            continue
        t = str(v).strip()
        if not t or t in vocab_values:
            continue
        return t
    return ""


def survey(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    setting = setting_columns(wb)
    vocab_values = {v for vals in setting.values() for v in vals}
    header_row, sample_cols = sample_table(ws)
    validated = validation_targets(ws)

    # The sample table sits to the RIGHT of the form block, so column — not row
    # — is what separates a per-sample dropdown from a form-level choice. Every
    # form puts its choices in columns A-G and its table from column H or I on.
    table_col = 99
    if sample_cols:
        table_col = openpyxl.utils.column_index_from_string(
            CELL_RE.match(sample_cols[0][0]).group(1))

    choices, per_sample = [], []
    for coord, src in sorted(validated.items(),
                             key=lambda kv: (int(CELL_RE.match(kv[0]).group(2)), kv[0])):
        m = CELL_RE.match(coord)
        col = openpyxl.utils.column_index_from_string(m.group(1))
        entry = {"cell": coord, "label": label_for(ws, coord, vocab_values), "source": src}
        if col >= table_col:
            per_sample.append(entry)
        else:
            choices.append(entry)

    # One cell per label is enough: forms merge controls across 2-3 cells.
    seen, deduped = set(), []
    for c in choices:
        key = c["label"] or c["cell"]
        if key in seen:
            continue
        seen.add(key)
        deduped.append(c)
    choices = deduped

    # Header/question labels: text cells in the left block that read as fields.
    header_fields = []
    for row in ws.iter_rows(max_row=min(ws.max_row, 40), max_col=8):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            t = str(v).strip()
            if not t or t.startswith("*") or len(t) > 90:
                continue
            if t.endswith(":") or t.endswith("?"):
                header_fields.append((cell.coordinate, t))

    return {
        "file": Path(path).name,
        "title": str(ws["A6"].value or "").strip(),
        "sheet": ws.title,
        "sheets": wb.sheetnames,
        "not_built": NOT_BUILT.get(Path(path).name),
        "header_fields": header_fields,
        "choice_fields": choices,
        "per_sample_dropdowns": per_sample,
        "sample_header_row": header_row,
        "sample_columns": [c[1] for c in sample_cols],
        "per_sample_sources": sorted({c["source"] for c in per_sample if c["source"]}),
        "setting": setting,
    }


def matrix(results):
    """Cross-application comparison — the variation is the point."""
    print("\n" + "=" * 78)
    print("SAMPLE-COLUMN MATRIX")
    print("=" * 78)
    all_cols = []
    for r in results:
        for c in r["sample_columns"]:
            if c not in all_cols:
                all_cols.append(c)
    for col in all_cols:
        users = [r["file"].split("-electronic")[0].split(".xlsx")[0][:22]
                 for r in results if col in r["sample_columns"]]
        print(f"  {col:<32} {len(users):>2}  {', '.join(users)}")

    print("\n" + "=" * 78)
    print("CHOICE-FIELD MATRIX")
    print("=" * 78)
    all_q = []
    for r in results:
        for c in r["choice_fields"]:
            lab = c["label"]
            if lab and lab not in all_q:
                all_q.append(lab)
    for q in all_q:
        users = [r["file"].split("-electronic")[0].split(".xlsx")[0][:22]
                 for r in results if any(c["label"] == q for c in r["choice_fields"])]
        print(f"  {q[:46]:<46} {len(users):>2}  {', '.join(users)}")

    print("\n" + "=" * 78)
    print("SETTING COLUMNS PER WORKBOOK")
    print("=" * 78)
    for r in results:
        print(f"  {r['file'][:44]:<44} {list(r['setting'].keys())}")


def markdown(results):
    """The per-application survey table the module prompt asks for."""
    out = ["# Submission-form field survey",
           "",
           f"Generated by `tools/survey_workbooks.py` from {len(results)} workbooks. "
           "Do not hand-edit — re-run the tool.",
           "",
           "## Per application",
           "",
           "| Application | Built? | Choice fields | Sample columns |",
           "|---|---|---|---|"]
    for r in results:
        name = r["file"].replace(".xlsx", "")
        built = f"no — {r['not_built']}" if r["not_built"] else "yes"
        ch = "<br>".join(c["label"][:44] for c in r["choice_fields"] if c["label"]) or "—"
        sc = "<br>".join(r["sample_columns"]) or "—"
        out.append(f"| `{name}` | {built} | {ch} | {sc} |")

    out += ["", "## Sample columns, by how many forms use them", "",
            "| Column | Forms | Used by |", "|---|---|---|"]
    all_cols = []
    for r in results:
        for c in r["sample_columns"]:
            if c not in all_cols:
                all_cols.append(c)
    for col in sorted(all_cols, key=lambda c: -sum(c in r["sample_columns"] for r in results)):
        users = [r["file"].split("-electronic")[0].split(".xlsx")[0]
                 for r in results if col in r["sample_columns"]]
        out.append(f"| `{col}` | {len(users)} | {', '.join(users)} |")

    out += ["", "## Setting vocabularies present", "",
            "| Workbook | Setting columns |", "|---|---|"]
    for r in results:
        out.append(f"| `{r['file']}` | {', '.join(r['setting'].keys()) or '—'} |")
    return "\n".join(out) + "\n"


def main():
    results = [survey(p) for p in workbooks()]

    if "--json" in sys.argv:
        print(json.dumps(results, ensure_ascii=False, indent=2))
        return

    if "--md" in sys.argv:
        dest = Path(__file__).resolve().parent.parent / "docs" / "FIELD_SURVEY.md"
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_text(markdown(results), encoding="utf-8")
        print(f"wrote {dest}")
        return

    for r in results:
        print("=" * 78)
        flag = f"   [NOT BUILT — {r['not_built']}]" if r["not_built"] else ""
        print(f"{r['file']}{flag}")
        print(f"  title : {r['title']}")
        print(f"  sheets: {r['sheets']}")
        print(f"  HEADER FIELDS ({len(r['header_fields'])}):")
        for coord, label in r["header_fields"]:
            print(f"    {coord:>5}  {label}")
        print(f"  CHOICE FIELDS ({len(r['choice_fields'])}):")
        for c in r["choice_fields"]:
            print(f"    {c['cell']:>5}  {c['label'][:52]:<52} <- {c['source']}")
        if r["per_sample_dropdowns"]:
            srcs = sorted({c["source"] for c in r["per_sample_dropdowns"]})
            print(f"  PER-SAMPLE DROPDOWNS: {len(r['per_sample_dropdowns'])} cells <- {srcs}")
        print(f"  SAMPLE COLUMNS (row {r['sample_header_row']}): {r['sample_columns']}")
        print(f"  SETTING COLUMNS: {list(r['setting'].keys())}")

    matrix(results)
    print("\n" + "=" * 78)
    print(f"{len(results)} workbooks surveyed")


if __name__ == "__main__":
    main()
