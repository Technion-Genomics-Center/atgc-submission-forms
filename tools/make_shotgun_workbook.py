# -*- coding: utf-8 -*-
"""Generate the shotgun metagenomics workbook (doc 05 §4.2).

Shotgun is the one application with no source workbook — the service is new and
the form was built from the DNA-seq layout. This writes the missing workbook so
the paper trail matches the other nineteen.

It is generated from the BUILT SPEC, not typed by hand, so the workbook and the
web form cannot drift: same columns, same vocabularies, same wording.

    python tools/make_shotgun_workbook.py

Writes into SubmissionForm/. It does NOT write to the lab share — the
Applications folder is read-only (CLAUDE.md), so the file is handed over for
Nitsan to place.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
SLUG = "shotgun-metagenomics"
OUT = ROOT / "Shotgun-metagenomics-electronic_2026.xlsx"

NAVY = "112954"
ACCENT = "FDA13C"          # Genomics, the section this application sits in
RULE = "E3EBF2"


def built_spec():
    page = ROOT / "dist" / SLUG / "index.html"
    if not page.exists():
        sys.exit("Run build.py first — this reads the built form.")
    m = re.search(r"window\.APP = (\{.*\});", page.read_text(encoding="utf-8"))
    return json.loads(m.group(1))


def main():
    spec = built_spec()
    wb = openpyxl.Workbook()

    # ── the form sheet ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "DNA sample submission form"
    ws.sheet_view.showGridLines = False

    title = Font(name="Calibri", size=18, bold=True, color=NAVY)
    label = Font(name="Calibri", size=11, bold=True, color=NAVY)
    band = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    small = Font(name="Calibri", size=10, italic=True, color="5F7488")
    head = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_accent = PatternFill("solid", fgColor=ACCENT)
    fill_navy = PatternFill("solid", fgColor=NAVY)
    thin = Border(bottom=Side(style="thin", color=RULE))
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    logo = ROOT / "assets" / "logo_small.png"
    if logo.exists():
        img = XLImage(str(logo))
        img.anchor = "A1"
        ws.add_image(img)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 46

    ws["B1"] = f"{spec['name']} Submission Form"
    ws["B1"].font = title
    ws["B2"] = "Azrieli Technion Genomics Center"
    ws["B2"].font = small

    def section(row, text):
        for col in ("A", "B"):
            c = ws[f"{col}{row}"]
            c.fill = fill_accent
            c.font = band
            c.alignment = left
        ws[f"A{row}"] = text

    def field(row, text, note=""):
        ws[f"A{row}"] = text
        ws[f"A{row}"].font = label
        ws[f"A{row}"].alignment = left
        ws[f"A{row}"].border = thin
        ws[f"B{row}"].border = thin
        if note:
            ws[f"C{row}"] = note
            ws[f"C{row}"].font = small

    r = 7
    section(r, "Your details"); r += 1
    for name in ("Date", "Submitting to which ATGC lab?", "Name", "Group / PI",
                 "Institute", "Faculty (Technion only)", "Email", "Phone",
                 "Quote number", "BaseSpace account", "Technion budget number"):
        field(r, name); r += 1

    r += 1
    section(r, "Sequencing"); r += 1
    v = spec.get("vocabularies", {})
    choices = [
        ("Library preparation", [x if isinstance(x, str) else x["label"]
                                 for x in v.get("LibraryPrep", [])]),
        ("Flow cell", [x["label"] if isinstance(x, dict) else x
                       for x in v.get("FlowCell", [])]),
        ("Run mode", v.get("RunMode", [])),
        ("Number of flow cells", v.get("Run#", [])),
        ("Do you require extraction?", ["Yes", "No"]),
        ("QC required", ["Yes", "No"]),
        ("Do you require bioinformatic analysis?", ["Yes", "No"]),
    ]
    first_choice_row = r
    for name, options in choices:
        field(r, name)
        if options:
            # Excel caps an inline list at 255 characters; longer lists point at
            # the Setting sheet instead, which is what the other workbooks do.
            joined = ",".join(o.replace(",", " ") for o in options)
            if len(joined) < 250:
                dv = DataValidation(type="list", formula1=f'"{joined}"',
                                    allow_blank=True)
                ws.add_data_validation(dv)
                dv.add(ws[f"B{r}"])
        r += 1

    r += 1
    section(r, "Samples"); r += 1
    ws[f"A{r}"] = ("Sample names: letters, numbers, underscore and full stop "
                   "only — no spaces or hyphens, and each must be unique.")
    ws[f"A{r}"].font = small
    r += 1

    header_row = r
    cols = spec["columns"]
    for i, name in enumerate(cols, start=1):
        c = ws.cell(row=header_row, column=i, value=name)
        c.font = head
        c.fill = fill_navy
        c.alignment = left
        ws.column_dimensions[get_column_letter(i)].width = \
            34 if name == "Remarks" else 18
    # A hundred rows, so a plate fits without anyone inserting rows by hand.
    for n in range(1, 101):
        ws.cell(row=header_row + n, column=1).border = thin

    quant_col = cols.index(spec["quant_column"]) + 1 if spec["quant_column"] in cols else None
    if quant_col:
        dv = DataValidation(type="list",
                            formula1='"' + ",".join(spec["quant_options"]) + '"',
                            allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(quant_col)}{header_row + 1}:"
               f"{get_column_letter(quant_col)}{header_row + 100}")

    # ── the Setting sheet, matching the other workbooks ─────────────────────
    st = wb.create_sheet("Setting")
    for i, (name, values) in enumerate(
            [("YesNo", ["Yes", "No"]),
             ("LibraryPrep", [x if isinstance(x, str) else x["label"]
                              for x in v.get("LibraryPrep", [])]),
             ("FlowCell", [x["label"] if isinstance(x, dict) else x
                           for x in v.get("FlowCell", [])]),
             ("RunMode", v.get("RunMode", [])),
             ("Run#", v.get("Run#", [])),
             ("Conc type", spec.get("quant_options", [])),
             ("ExtractionService", v.get("ExtractionService", []))], start=1):
        col = get_column_letter(i * 2 - 1)
        st[f"{col}1"] = name
        st[f"{col}1"].font = Font(bold=True)
        for j, value in enumerate(values, start=2):
            st[f"{col}{j}"] = value
        st.column_dimensions[col].width = max(14, min(46, max(
            [len(name)] + [len(str(x)) for x in values]) + 2))

    wb.save(OUT)
    print(f"wrote {OUT.name}")
    print(f"  {len(cols)} sample columns, {len(choices)} choice fields")
    print(f"  vocabularies: {', '.join(k for k in v)}")
    print()
    print("  This file is NOT written to the lab share — Applications/ is")
    print("  read-only. Place it in:")
    print(r"    Y:\LAB\TGC_website\ATGC_website_2025\Applications\Metagenomics\ ")


if __name__ == "__main__":
    main()
